# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class WalkinReport(models.TransientModel):
    _name = 'walkin.report'
    _description = 'Walkin Report'

    institution = fields.Many2one('sh.medical.health.center', string='Health Center', help="Health Center", required=True,
                                  default=lambda self: self.env.ref('shealth_all_in_one.sh_medicines_health_center_kangnam_hn'))
    report_type = fields.Selection([('walkin_by_room', 'BC khám bệnh theo phòng'), ('walkin_by_patient', 'BC khám bệnh theo bệnh nhân'),
                                    ('surgery', 'BC phẫu thuật'), ('lab_image', 'BC cận lâm sàng'), ('walkin_by_exam_dep_specialty', 'BC hoạt động khám bệnh chuyên khoa')],
                                   string='Report type', default='walkin_by_room')
    month = fields.Selection([(1, '1'), (2, '2'), (3, '3'), (4, '4'), (5, '5'), (6, '6'), (7, '7'), (8, '8'), (9, '9'), (10, '10'), (11, '11'), (12, '12')],
                             string='Month', default=date.today().month)
    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def _get_range_walkin_data(self):
        #lay khoa kham benh cua benh vien
        exam_dep = self.env['sh.medical.health.center.ward'].search([('institution', '=', self.institution.id), ('type', '=', 'Examination')], limit=1)
        # lay khoa kham benh cua benh vien
        all_exam_room = self.env['sh.medical.health.center.ot'].search([('department','=',exam_dep.id)])
        data = []
        for room in all_exam_room:
            room_walkin = self.env['sh.medical.appointment.register.walkin'].search_count([('service_room', '=', room.id),
                                                                                           ('state', 'not in', ['Scheduled', 'WaitPayment']),  # Thanh mới thêm vào
                                                                                           ('date', '>=', self.start_date),
                                                                                           ('date', '<=', self.end_date)])
            data.append({'code': room.code,'name': room.name,'total': room_walkin})

        return data
    #SonDC
    def report_walkin(self):
        walkin_attachment = self.env['ir.attachment'].browse(self.env.ref('shealth_all_in_one.walkin_report_attachment').id)
        decode = base64.b64decode(walkin_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['e3'].value = 'Từ ngày: %s đến ngày: %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        data = self._get_range_walkin_data()
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)
        key_col_list = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
        key_list = [
            'code',
            'name',
            'total',
            'SLK_YHCT',
            'SLK_TE',
            'SLK_BHYT',
            'SLK_VP',
            'SLK_KTP',
            'SLK_CC',
            'SNBVV',
            'SNCV',
            'DTNT_SNB',
            'DTNT_YHCT',
            'DTNT_TE',
            'DTNT_SN',
        ]
        row = 7
        for line_data in data:
            ws.cell(row, 2).value = row - 6
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value = line_data.get(k)
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'MS report',
            'datas_fname': 'BAO_CAO_HOAT_DONG_KHAM_BENH.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })

        return {
            'name': 'Báo cáo hoạt động khám bệnh theo phòng',
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'attachment_id': attachment.id}
        }


    # báo cáo khám bệnh theo phòng chia sheet
    # def report_walkin(self):
    #     walkin_attachment = self.env['ir.attachment'].browse(self.env.ref('shealth_all_in_one.walkin_report_attachment').id)
    #     decode = base64.b64decode(walkin_attachment.datas)
    #     wb = load_workbook(BytesIO(decode))
    #     temp_ws = wb.active
    #     temp_ws['e3'].value = 'Từ ngày: %s đến ngày: %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
    #     temp_ws['b2'].value = str.upper(self.institution.name)
    #     data = self._get_range_walkin_data()
    #     wses = [temp_ws]
    #     if len(data) > 1:
    #         temp_ws.title = data[0]['code']
    #         for i in range(1, len(data)):
    #             new_ws = wb.copy_worksheet(temp_ws)
    #             new_ws.title = data[i]['code']
    #             wses.append(new_ws)
    #     for item, ws in zip(data, wses):
    #         row = 7
    #         for col in range(2, 18):
    #             ws.cell(row=row, column=col).border = all_border_thin
    #             ws.cell(row=row, column=col).font = Font(name='Times New Roman', size=12)
    #             index = row - 6
    #             keys = ['code', 'name', 'total']
    #             for col, key in enumerate(keys, 3):
    #                 ws.cell(row=row, column=col).value = item[key]
    #             ws.cell(row=row, column=2).value = index
    #             ws.cell(row=row, column=9).value = ws.cell(row=row, column=14).value = item['total']
    #         row += 1
    #     # ws.cell(row=row+2, column=9).value = 'Ngày %s tháng %s năm %s ' % (fields.date.today().strftime('%d'), fields.date.today().strftime('%m'), fields.date.today().strftime('%Y'))
    #         sign1_col, sign2_col, sign3_col = 4, 9, 14
    #         if 'PTTT' in item['code']:  # Todo: check this condition later
    #             sign1_col, sign2_col, sign3_col = 3, 5, 12
    #             ws.unmerge_cells('E1:Q2')
    #             ws.unmerge_cells('E3:Q3')
    #             ws.delete_cols(14, 4)
    #             ws.merge_cells('E1:M2')
    #             ws.merge_cells('E3:M3')
    #             ws.print_area = "A1:N18"
    #         ws.cell(row=row+2, column=sign3_col).value = 'Ngày ... tháng ... năm ... '
    #         ws.cell(row=row+2, column=sign3_col).font = Font(name='Times New Roman', italic=True, size=10)
    #         ws.cell(row=row+2, column=sign3_col).alignment = Alignment(horizontal='center')
    #
    #         ws.cell(row=row+3, column=sign1_col).value, ws.cell(row=row+4, column=sign1_col).value = 'NGƯỜI LẬP BIỂU ', '(Chức danh, Ký tên)'
    #         ws.cell(row=row + 3, column=sign1_col).alignment = ws.cell(row=row + 4, column=sign1_col).alignment = Alignment(horizontal='center')
    #         ws.cell(row=row + 3, column=sign1_col).font = Font(name='Times New Roman', size=10, bold=True)
    #         ws.cell(row=row + 4, column=sign1_col).font = Font(name='Times New Roman', size=10, italic=True)
    #
    #         ws.cell(row=row + 3, column=sign2_col).value, ws.cell(row=row + 4, column=sign2_col).value = 'TRƯỞNG PHÒNG KHTH ', '(Chức danh, Ký tên)'
    #         ws.cell(row=row + 3, column=sign2_col).alignment = ws.cell(row=row + 4, column=sign2_col).alignment = Alignment(horizontal='center')
    #         ws.cell(row=row + 3, column=sign2_col).font = Font(name='Times New Roman', size=10, bold=True)
    #         ws.cell(row=row + 4, column=sign2_col).font = Font(name='Times New Roman', size=10, italic=True)
    #
    #         ws.cell(row=row + 3, column=sign3_col).value, ws.cell(row=row + 4, column=sign3_col).value = 'GIÁM ĐỐC ', '(Chức danh, Ký tên)'
    #         ws.cell(row=row + 3, column=sign3_col).alignment = ws.cell(row=row + 4, column=sign3_col).alignment = Alignment(horizontal='center')
    #         ws.cell(row=row + 3, column=sign3_col).font = Font(name='Times New Roman', size=10, bold=True)
    #         ws.cell(row=row + 4, column=sign3_col).font = Font(name='Times New Roman', size=10, italic=True)
    #
    #     fp = BytesIO()
    #     wb.save(fp)
    #     fp.seek(0)
    #     report = base64.encodebytes((fp.read()))
    #     fp.close()
    #     attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
    #                                                          'datas_fname': 'BAO_CAO_HOAT_DONG_KHAM_BENH.xlsx',
    #                                                           'datas': report,
    #                                                           'res_model': 'temp.creation',
    #                                                           'public': True})
    #     return {'name': 'BÁO CÁO HOẠT ĐỘNG KHÁM BỆNH',
    #             'type': 'ir.actions.act_window',
    #             'res_model': 'temp.wizard',
    #             'view_mode': 'form',
    #             'view_type': 'form',
    #             'target': 'inline',
    #             'view_id': self.env.ref('ms_templates.report_wizard').id,
    #             'context': {'attachment_id': attachment.id}}
    #
    #     url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=BAO_CAO_HOAT_DONG_KHAM_BENH.xlsx" \
    #           % (attachment.id)
    #     cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
    #     cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
    #     return {'name': 'BÁO CÁO HOẠT ĐỘNG KHÁM BỆNH',
    #             'type': 'ir.actions.act_url',
    #             'url': url,
    #             'target': 'self',
    #             }

    def report_walkin_patient(self):
        template = self.env.ref('shealth_all_in_one.bao_cao_kham_benh')
        walkins = self.env['sh.medical.appointment.register.walkin'].search([('institution', '=', self.institution.id),
                                                                            ('state', 'not in', ['Scheduled', 'WaitPayment']),
                                                                            ('date', '>=', self.start_date),
                                                                            ('date', '<=', self.end_date)])
        return {'name': (_('Báo cáo khám bệnh')),
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'default_template_id': template.id, 'active_ids': walkins.ids,
                            'active_model': 'sh.medical.appointment.register.walkin'}}

    def report_surgery(self):
        template = self.env.ref('shealth_all_in_one.bao_cao_phau_thuat')
        start_date = self.start_date
        end_date = self.end_date
        if start_date.month != end_date.month:
            raise UserError(_('Nhập ngày bắt đầu và kết thúc trong cùng một tháng'))
        if start_date.day != 1:
            raise UserError(_('Nhập ngày bắt đầu là mùng 1'))
        surgeries = self.env['sh.medical.surgery'].search([('institution', '=', self.institution.id),
                                                          ('surgery_date', '>=', start_date),
                                                          ('surgery_date', '<=', end_date)])
        previous_surgery_count = 0
        if start_date.month > 1:
            previous_surgery_count = self.env['sh.medical.surgery'].search_count([('institution', '=', self.institution.id),
                                                                                  ('state', '!=', 'Draft'),
                                                                                  ('surgery_date', '>=', date(start_date.year, 1, 1)),
                                                                                  ('surgery_date', '<', start_date)])
        external_keys = {'a2': start_date.strftime('DANH SÁCH KHÁCH HÀNG KHOA PHẪU THUẬT THẨM MỸ THÁNG %m/%Y'),
                         'b6': previous_surgery_count + 1}

        return {'name': (_('Báo cáo phẫu thuật')),
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'default_template_id': template.id, 'active_ids': surgeries.ids, 'external_keys': external_keys,
                            'active_model': 'sh.medical.surgery'}}

    def report_lab_imaging(self):
        lab_imaging_attachment = self.env['ir.attachment'].browse(self.env.ref('shealth_all_in_one.lab_imaging_report_attachment').id)
        decode = base64.b64decode(lab_imaging_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        title_font = Font(name='Times New Roman', size=12, bold=True, italic=True)

        lab = self.env['sh.medical.lab.test']
        imaging = self.env['sh.medical.imaging']
        lab_group_types = self.env['sh.medical.lab.group.type'].search([])
        imaging_cdha = self.env['sh.medical.imaging.group.type'].search([('type', '=', 'cdha')])
        imaging_tdcn = self.env['sh.medical.imaging.group.type'].search([('type', '=', 'tdcn')])

        ws['a3'].value = self.end_date.strftime('Tháng %m Năm %Y')
        ws['g54'].value = datetime.now().strftime('Ngày %d Tháng %m Năm %Y')

        row = 7
        for lab_group_type in lab_group_types:
            ws.cell(row, 1).value = lab_group_type.name
            print(lab_group_type.name)
            ws.cell(row, 2).value, ws.cell(row, 2).border = 'Xét nghiệm', all_border_thin
            ws.cell(row, 3).value = len(lab.read_group(domain=[('date_done', '>=', self.start_date),
                                                               ('date_done', '<=', self.end_date),
                                                               ('group_type', '=', lab_group_type.id),
                                                               ('state', '=', 'Completed')],
                                                       fields=['id'], groupby=['walkin']))
            row += 1

        row += 1
        ws.cell(row, 1).value, ws.cell(row, 1).font = 'II. Chẩn đoán hình ảnh', title_font
        row += 1
        for image in imaging_cdha:
            ws.cell(row, 1).value = image.name
            ws.cell(row, 2).value = 'Lần'
            ws.cell(row, 3).value = len(imaging.read_group(domain=[('date_done', '>=', self.start_date),
                                                                   ('date_done', '<=', self.end_date),
                                                                   ('group_type', '=', image.id),
                                                                   ('state', '=', 'Completed')],
                                                           fields=['id'], groupby=['walkin']))
            row += 1

        last_row = row
        row = 7
        for image_cn in imaging_tdcn:
            ws.cell(row, 6).value = image_cn.name
            ws.cell(row, 7).value = 'Lần'
            ws.cell(row, 8).value = len(imaging.read_group(domain=[('date_done', '>=', self.start_date),
                                                                   ('date_done', '<=', self.end_date),
                                                                   ('group_type', '=', image_cn.id),
                                                                   ('state', '=', 'Completed')],
                                                           fields=['id'], groupby=['walkin']))
            row += 1

        row += 1
        ws.cell(row, 6).value, ws.cell(row, 6).font = 'IV. Truyền máu:', title_font
        row += 1
        ws.cell(row, 6).value, ws.cell(row, 7).value = 'Số ml truyền máu', 'lít'
        row += 2
        ws.cell(row, 6).value, ws.cell(row, 6).font = 'V. Giải phẫu bệnh', title_font
        row += 1
        for i in ['Đại thể', 'Vi thể', 'Khác']:
            ws.cell(row, 6).value, ws.cell(row, 7).value = i, 'Mẫu'
            row += 1

        last_row = row if row > last_row else last_row
        ws.delete_rows(last_row, 48-last_row)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                              'datas_fname': 'BAO_CAO_CAN_LAM_SANG.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})
        return {'name': 'BÁO CÁO CẬN LÂM SÀNG',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}


    # SonDC
    def report_walkin_exam_dep_specialty(self):
        # search các khoa kham benh chuyen khoa
        exam_dep = self.env['sh.medical.health.center.ward'].search(
            [('institution', '=', self.institution.id), ('type', '=', 'Specialty')])
        data = []
        for i in exam_dep:
            exam_dep_walkin = self.env['sh.medical.appointment.register.walkin'].search_count(
                [('service_room.department', '=', i.id),
                 ('state', 'not in', ['Scheduled', 'WaitPayment']),
                 ('date', '>=', self.start_date),
                 ('date', '<=', self.end_date),
                 ])
            data.append({'name': i.name, 'total': exam_dep_walkin,})

        walkin_attachment = self.env['ir.attachment'].browse(self.env.ref('shealth_all_in_one.walkin_report_exam_dep_spec_attachment').id)
        decode = base64.b64decode(walkin_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        ws['a3'].value = 'Từ ngày: %s đến ngày: %s' % (self.start_date.strftime('%d/%m/%Y'), self.end_date.strftime('%d/%m/%Y'))
        ws['k35'].value = date.today().strftime('Ngày %d tháng %m năm %Y')
        for i in data:
            print(i)
            if i['name'] == 'Khoa Da liễu':
                ws.cell(21,3).value = i['total']
            if i['name'] == 'Phòng Răng - Hàm - Mặt':
                ws.cell(32,3).value = i['total']

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'MS report',
            'datas_fname': 'BAO_CAO_HOAT_DONG_KHAM_BENH_CHUYEN_KHOA.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })

        return {
            'name': 'Báo cáo hoạt động khám bệnh chuyên khoa',
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'attachment_id': attachment.id}
        }